import openpyxl

"""
엑셀에 저장
1. 파일 생성
2. 시트 생성
"""

"""
빈파일 = openpyxl.Workbook()

# 추가로 엑셀 파일 불러오는 방법
파일 = openpyxl.load_workbook('경로')

# 단순 생성 (시트명은 sheet로 생성됨)
시트 = 빈파일.active

# 시트명 명시를 통한 생성
시트 = 빈파일.create_sheet('시트명')

# 시트의 특정 셀에 저장
시트.cell(row = 행, column = 열).value

# 파일 확장자 포함. xlsx
파일.save('경로.xlsx')

# 파일의 시트 목록
시트목록 = 파일.sheetnames

# 특정 시트 선택
시트 = 엑셀파일['시트명']
"""

# 엑셀파일 생성
xlsx_file = openpyxl.Workbook()

# 생성한 파일에 시트 생성
xlsx_sheet = xlsx_file.active

# 시트 특정 셀에 데이터 입력

for i in range(10):
    xlsx_sheet.cell(row=i + 1, column=1).value = "hi"
# value에 find_element().text로 찾은 값을 넣으면 됨


# 저장
xlsx_file.save("result.xlsx")
