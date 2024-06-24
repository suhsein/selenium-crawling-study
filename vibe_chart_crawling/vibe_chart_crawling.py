# 셀레늄 웹드라이버
from selenium import webdriver

# 옵션
from selenium.webdriver.chrome.options import Options

# 선택자 및 키보드 입력
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# sleep
import time

# 엑셀 저장
import openpyxl
######################################################################

# 드라이버 생성
custom_options = Options()
# 헤드리스 모드로 동작
custom_options.add_argument('--headless')

browser = webdriver.Chrome(options=custom_options)
browser.maximize_window()

# URL로 웹페이지 가져오고 렌더링 완료까지 기다림
URL = "https://vibe.naver.com/chart/total"
browser.get(URL)
browser.implicitly_wait(10)

######################################################################

# 팝업창 닫기
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/a[2]').click()

# 엑셀 파일, 시트 만들기
## xlsx_file = openpyxl.Workbook()
## xlsx_sheet = xlsx_file.active

# 긁어오기
for i in range(1, 101, 1):
    path = f'//*[@id="content"]/div[4]/div[2]/div/table/tbody/tr[{i}]'
    rank = browser.find_element(By.XPATH, f'{path}/td[3]/span').text # 순위
    title = browser.find_element(By.XPATH, f'{path}/td[4]/div[1]/span/a').text # 곡명
    artist = browser.find_element(By.XPATH, f'{path}/td[5]/span/span/span/a/span').text # 가수명
    
    # sheet 셀에 저장
    ## xlsx_sheet.cell(row=i, column= 1).value = rank
    ## xlsx_sheet.cell(row=i, column= 2).value = title
    ## xlsx_sheet.cell(row=i, column= 3).value = artist
    
    ## print(rank, title, artist)

# 엑셀 파일 저장
## xlsx_file.save('vibe_top100.xlsx')