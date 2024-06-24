# 웹드라이버
from selenium import webdriver

# 옵션
from selenium.webdriver.chrome.options import Options

# 선택자, 키보드 조작
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import time

# 엑셀 저장
import openpyxl

###############################################################

# 웹드라이버 세팅
CustomOption = Options()
browser = webdriver.Chrome(options=CustomOption)
browser.maximize_window


# URL로 웹페이지 읽기
URL = "https://datalab.naver.com/shoppingInsight/sCategory.naver"
browser.get(URL)
browser.implicitly_wait(10)

"""
분야 선택

분야 1, 분야 2, 분야 3 존재
분야 1은 필수선택

1. 드롭다운 메뉴 목록 보기 버튼 클릭
2. 파이썬 변수가 가지는 값과 동일한 text를 가지는 요소 클릭

contains() 사용 -> 첫번째 파라미터로 text(), 두번째 파라미터로 value. 단 value를 포함하는 첫번째 요소만 반환
text() 사용 -> text() = value. value와 완전히 같은 요소를 반환
"""

class1 = "화장품/미용"
class2 = "향수"
class3 = "여성향수"
class4 = ""

select_box_path = '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div'
select_menu_path = '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div'

# 분야 1
browser.find_element(By.XPATH, f"{select_box_path}[1]/span").click()
browser.find_element(
    By.XPATH, f"{select_menu_path}[1]/ul/li/a[text()='{class1}']"
).click()

# 분야 2
if class2 != "":
    browser.find_element(By.XPATH, f"{select_box_path}[2]/span").click()
    browser.find_element(
        By.XPATH, f"{select_menu_path}[2]/ul/li/a[text()='{class2}']"
    ).click()

# 분야 3
if class3 != "":
    browser.find_element(By.XPATH, f"{select_box_path}[3]/span").click()
    browser.find_element(
        By.XPATH, f"{select_menu_path}[3]/ul/li/a[text()='{class3}']"
    ).click()

# 조회하기
browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/a').click()
time.sleep(3)

# 엑셀 파일, 시트 생성
xlsx_file = openpyxl.Workbook()
xlsx_sheet = xlsx_file.active

# 총 25페이지
for i in range(25):
    for j in range(1, 21, 1):
        path = (
            f'//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{j}]/a'
        )
        # 긁어오기
        keyword = browser.find_element(By.XPATH, f"{path}").text.split('\n')

        xlsx_sheet.cell(row=i * 20 + j, column=1).value = keyword[0]
        xlsx_sheet.cell(row=i * 20 + j, column=2).value = keyword[1]

    # 버튼 누르기
    browser.find_element(
        By.XPATH, '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]'
    ).click()
    time.sleep(0.5)


xlsx_file.save("naver_shopping_top500.xlsx")
